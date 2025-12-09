# 打印日历
import calendar
cal = calendar.month(2020, 3)
print(cal)

# -----------------------------------------------------------------
def Text_1(*parama):
    for i in parama:
        print(f"r{i}")

you = input("1:")
are = input("2:")
Text_1(you, are) # 不定长参数

# -----------------------------------------------------------------
get_answer = lambda c, e, n: c + int(e / n)
print(get_answer(9, 15, 3))

# -----------------------------------------------------------------
class Mylover:
    def __init__(self, name, age):
        self.name = name
        self.age = age
    def find(self):
        print(f"{self.name}")
    def __del__(self):
        print(f"{self}销毁")
n = Mylover("lijunyan", 20)
n.find()
print (f"MyMylover.__doc__{Mylover.__doc__}")
print (f"MyMylover.__name__{Mylover.__name__}")
print (f"MyMylover.__module__:{Mylover.__module__}")
print (f"MyMylover.__bases__:{Mylover.__bases__}")
print (f"MyMylover.__dict__:{Mylover.__dict__}")
del n

# -----------------------------------------------------------------
import math
print(dir(math))

# -----------------------------------------------------------------
myfile_ = open("玫瑰", "r", encoding="UTF-8")
# myfile_.write("我心底有一簇向着烈阳而升的花\n比一切美酒都要芬芳\n滚烫的馨香\n淹没过稻草人的胸膛\n草扎的精神\n从此万寿无疆\n")
position = myfile_.tell()
print(position)
position = myfile_.seek(16, 0)
print(position)
# str_ = myfile_.read(4)
# print(str_)
myfile_.close()

# -----------------------------------------------------------------
import os
# os.rename("玫瑰", "花")
open("NULL", "a")
os.remove("NULL")
os.mkdir("TEXT_10_8_1")
os.chdir("./AVS_PYTHON/TEXT_10_8_1")
print(os.getcwd())
os.rmdir("TEXT_10_8_1")

# -----------------------------------------------------------------
import re
re.match

# -----------------------------------------------------------------
import sys
newpath = "d:\\PPioneer\\IDFC\\AVS_CITY\\AVS_PYTHON"
sys.path.append(newpath)
print(sys.path)
# 修改路径

# -----------------------------------------------------------------
# 设置动态路径
print(__file__) #例：将文件的路径改为相对路经，然后通过它进行展示
import os
import sys
dynamicpath = os.path.dirname(os.path.dirname(__file__))
print(dynamicpath)
sys.path.append(dynamicpath)
print(sys.path)
import os
print(os.getcwd())

# -----------------------------------------------------------------
from openpyxl import Workbook, load_workbook
wb = Workbook()
sheet = wb.active
if sheet is not None:
    sheet.title = "accept soluble"
    # print(sheet.title)
    sheet["A1"] = "a name"
    sheet["B2"] = "a age"
    sheet.append(["最下行", "最左边", "可多行"])
    wb.save("text_e.xlsx")  # 修复文件扩展名

# -----------------------------------------------------------------
try:
    wbanother = load_workbook("text_e.xlsx")
except FileNotFoundError:
    print("文件 text_e.xlsx 未找到")
except Exception as e:
    print(f"加载工作簿时出错: {e}")

# -----------------------------------------------------------------
subs_a = bytes("找到它", "UTF-8") # bytes:encoding->指定字节串的编译格式
subs_b = b"YOUNG"
#字节串的两种初始化方式->b"" & bytes()
print(f"1.{subs_a} and {subs_b}")
print(f"1-1.{ord("等待")}") # 字符转unicode字节
print(f"2.{chr(121)}") # unicode字节转字符
print(f"3.{subs_b[2]}") #索引
print(f"4.{subs_a.decode("UTF-8")}") # self.decode->字节串转字符串;
print(f"5.{"秋天".encode("UTF-8")}") #self.encode->字符串转字节串;
subs_c = bytearray("布散烈烈朝晖之时", "UTF-8") # 字节数组 bytearray();
print(f"6.{subs_c}")
subs_c.append(121)
print(f"7.{subs_c}")
print(f"8.{subs_c.decode("UTF-8")}")

# -----------------------------------------------------------------
subs_a = {"a", "b", "c", "d"}
print(f"1.{subs_a}") #集合->本身为可变，但其元素不可变,且不重复;
subs_b = frozenset({"a", "b"}) #冻结集合->本身也不可变;
print(f"2.{subs_b}")

# -----------------------------------------------------------------
# 整数、浮点数、布尔->按值分配地址
# 复数、列表、元组、集合、自定义对象->存储相应变量对应的元素的地址
# 字符串->对象的地址，若字符串相同，则为同一个地址

# -----------------------------------------------------------------
subs_a = 18.8;
print(f"1.{int(subs_a)}")
print(f"2.{int("444", 16)}") #base进制的数转为十进制，该数为字符串且不能为小数
print(f"3.{float(subs_a)}")

# -----------------------------------------------------------------
subs_a = eval("6 + 1") #eval参数必须是字符串
print(f"1.{subs_a}")

# -----------------------------------------------------------------
# //->取整商 **->幂 /->真除

# -----------------------------------------------------------------
subs_a = round(1.34655, 3)
print(f"1.{subs_a}")

# -----------------------------------------------------------------
print("younugug\\" #\->隔行符
"ja;sdjfa\na;\\"
"ald")

# -----------------------------------------------------------------
subs_a = 10
subs_b = 11 if subs_a < 1 else 0 # 三元运算
print(subs_b)

# -----------------------------------------------------------------
subs_a = 10
match subs_a:
    case 69 | 80:
        print(f"1.{subs_a}")
    case subs_b if subs_b < 69: # 使用变量接收match后面的表达式值，并可以使用if语句添加判断条件
        print("2.subs_b:", subs_b + 10)
    case _:  # 默认分支
        print(f"3.{None}")

# -----------------------------------------------------------------
subs_c = [1, 2, "text"]
match subs_c:
    case [1, 2, 3]:
        print(f"1.{subs_c}")
    case [1, (2 | 3), 3]:
        print(f"2.{subs_c}")


    case [(1 | 10), y]: 
        print(f"3.{y}")
    case [1, (3 | 2) as n, y] if len(y) < 2: #使用变量接受表达式的值，并可使用if语句添加判断条件
        print("4.n =", n, "y =", y)
    case [1, *rest]: #使用*变量接收剩下的元素，并生成列表
        print("5.rest =", rest)
    case _:
        print(f"3.{subs_c}")

# -----------------------------------------------------------------
subs_d = {'first' : "warrior", 'second' : 2}
match subs_d:
    case {'firsts' : _}:
        print("key:first存在")
    case {'firsts' : x, 'seconds' : y}:
        print("first:", x, 'second:', y)
    case {'first' : x, **rest}:
        print("rest =", rest)

# -----------------------------------------------------------------
subs = enumerate(['first', 'second', 'third'], 1)
for i in range(3):
    print(next(subs), end = ' ')

# -----------------------------------------------------------------
subs_zip = ['first', 'second', 'third']
subs_z = dict(zip(['you', 'he', 'I'], subs_zip), strict = True)
print(subs_z)

# -----------------------------------------------------------------
import secrets, string

def random_string(length) -> str:
    chars = string.ascii_letters + string.digits + string.punctuation
    return "".join(secrets.choice(chars) for i in range(length))

print(random_string(100))
